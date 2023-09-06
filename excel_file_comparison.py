import sys

# Library to manipulate excel
import openpyxl
from openpyxl.styles import PatternFill


class CompareExcel:
    """Class to compare two excel files"""

    def __init__(self, file1: str, file2: str):
        self.file1 = file1
        self.file2 = file2

    def generate_comparison_report(self):
        """Generate Comparison Report:This method two Excel cell by cell"""

        # Output file names
        output_file = "C:\\Folder\\ComparisonFile.xlsx"

        # Load excels
        wb1 = openpyxl.load_workbook(self.file1)
        wb2 = openpyxl.load_workbook(self.file2)


        # Pattern to fill (in this case red is taken)
        fill_pattern_red = PatternFill(patternType="solid", fgColor="FF3333")

        # number of sheets in each excel
        no_of_sheets1 = len(wb1.sheetnames)
        no_of_sheets2 = len(wb2.sheetnames)
        print(wb1.sheetnames)
        print(wb2.sheetnames)

        # Compare number of sheets
        if no_of_sheets1 != no_of_sheets2:
            print("Number of sheets are different in both the workbook")
            sys.exit(
                "noOfSheets1 : "
                + str(no_of_sheets1)
                + " noOfSheets2 : "
                + str(no_of_sheets2)
            )

        # If number of sheet is equal in both
        for s in range(0, no_of_sheets1):
            mismatchfound = 0
            sh1 = wb1.worksheets[s]
            sh2 = wb2.worksheets[s]

            sheet_name = wb1.sheetnames[s]

            row1 = sh1.max_row
            row2 = sh2.max_row
            row_max = max(row1, row2)
            # Compare number of rows
            if row1 != row2:
                print(
                    "Number of rows are different in both the sheet for : "
                    + sheet_name
                    + "\n"
                )

            column1 = sh1.max_column
            column2 = sh2.max_column
            column_max = max(column1, column2)

            # Compare number of rows
            if column1 != column2:
                print(
                    "Number of columns are different in both the sheet for : "
                    + sheet_name
                    + "\n"
                )

            extra_row = False
            extra_column = False

            for r in range(1, row_max + 1):
                for c in range(1, column_max + 1):
                    value1 = sh1.cell(r, c).value
                    value2 = sh2.cell(r, c).value
                    # Coloring extra columns and rows
                    if r > row1 or r > row2:
                        if not extra_row:
                            pass
                        sh2.cell(r, c).fill = fill_pattern_red
                        extra_row = True
                    elif c > column1 or c > column2:
                        if not extra_column:
                            pass
                        sh2.cell(r, c).fill = fill_pattern_red
                        extra_column = True
                    # Comparing both cells value
                    elif value1 == value2:
                        pass
                    else:
                        sh2.cell(r, c).fill = fill_pattern_red
                        mismatchfound += 1
        wb2.save(output_file)
        print("Comparison Done")


if __name__ == "__main__":
    """Main Function"""

    file1 = "C:\\Folder\\File1.xlsx"
    file2 = "C:\\Folder\\File2.xlsx"
    CompareExcel(file1, file2).generate_comparison_report()
